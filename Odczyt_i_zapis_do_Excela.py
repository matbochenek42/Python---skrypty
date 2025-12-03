import openpyxl as xl
from pathlib import Path
import time 

ktory_komputer = Path(r"Nazwa") #onedrive na różnych komputerach

#ścieżki do plików
sciezka_pliku = ktory_komputer / r"Nazwa.xlsx"
do_pliku = ktory_komputer / r"Nazwa.txt"
zapis_do_excela = ktory_komputer / r"Nazwa.xlsx"


def sprawdzanie_pliku():
    #przypisanie prawidłowej ścieżki do pliku, nazw pliku i arkusza
    nazwa_arkusza = "nazwa"
    
    try:
        wb = xl.load_workbook(sciezka_pliku, data_only=True)
        return wb, wb[nazwa_arkusza]
    except Exception:
        print(f"\nBłąd! Nie można otworzyć pliku")
        raise SystemExit

def funkcja_1(start = 4, koniec = 67, pozycja = 1, kol_1 = 5, kol_2 = 6):  #funkcja odczytująca zapisy z pliku excel
    
    wb, arkusz = sprawdzanie_pliku()

    lista = [ #zapis z kolumn do listy
        [arkusz.cell(row, pozycja).value,
        arkusz.cell(row, kol_1).value,
        arkusz.cell(row, kol_2).value]
        for row in range(start, koniec)
    ]

    wb.close()
    
    #odpowiednia zmiana poszczególnych wartości w kolumnach (niektóre komórki są scalowane albo zawierają puste wartości)
    lista[21][0] = None
    lista[22][0] = "2.2.8" 
    lista[23][0] = "2.2.8"

    lista[47][0] = None
    lista[48][0] = "2.2.23" 
    lista[49][0] = "2.2.23"

    lista[60][0] = None
    lista[61][0] = "2.2.32" 
    lista[62][0] = "2.2.32"

    return lista

def rozbieznosci(nazwa): #sprawdzanie rozbieżnosći ---> jeśli są zapis do listy
    lista = []

    for pozycja, kol1, kol2 in nazwa:
        if kol1 is not None and kol2 is not None and kol1 != kol2:
            
            roznica = kol1 - kol2
            
            if roznica > 0:
                porownanie = "mniejsza"
            elif roznica < 0:
                porownanie = "większa"

            lista.append(
                f".... {pozycja} ... {porownanie} o {abs(roznica)} ..."
            )
    return lista


# switch case

while True:
    print("""\n
========MENU=============
1. Dopisz do pliku
2. Zresetuj zawartość pliku
3. Wyjdź z programu
==========================\n""")  

    x = input("Napisz 1, 2 lub 3 w zależności od tego co chcesz zrobić - zgodnie z menu: ")

    match x: #zczytanie pozycji z listy i dodanie ich do pliku tekstowego (do wglądu), a później automatycznie z listy do excela 
        case "1":
            
            nazwa = funkcja_1()
            rozb = rozbieznosci(nazwa)

            if not rozb:
                print("\n Brak rozbieżności")
            else:
                with open(do_pliku, 'a', encoding = "utf-8") as f: #zapis do piku tekstowego
                    for _ in rozb:
                        f.write(f"{_}\n")
   
                
                    wb, arkusz = sprawdzanie_pliku() #bezpośredni zapis do excela w celu ostatecznego wykorzystania

                    for i, wiersz in enumerate(rozb):
                        arkusz.cell(row = 68 + i, column = 1, value = wiersz)

                    wb.save(zapis_do_excela)
                    wb.close()
                    
                print("\n✅ Udało się dopisać do pliku.")

        case "2": #zresetowanie wszystkich zapisów do 0
            
            with open(do_pliku, 'w', encoding = "utf-8") as f:
                f.write("")
            
            rozb = []

            wb, arkusz = sprawdzanie_pliku() #reset w excelu
            for wiersz in range(68,87):
                if wiersz is not None:
                    arkusz.cell(wiersz,1).value = ""

            wb.save(zapis_do_excela)
            wb.close()

            print("\n✅ Zresetowano docelowy plik. Można dodać dane na kolejny miesiąc")
            
        case "3": #wyjście z programu
            print("\nWyjdź z programu")
            time.sleep(1)
            break
        
        case _: #błędne wybranie opcji
            print("\n❌ Błąd! Wybrałeś opcję spoza menu")


